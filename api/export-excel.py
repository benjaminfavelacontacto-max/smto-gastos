from http.server import BaseHTTPRequestHandler
import json, os, sys, tempfile, traceback, re
from datetime import datetime

IMPORT_ERROR = None
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.workbook.protection import WorkbookProtection
    from PIL import Image as PILImage
except Exception:
    IMPORT_ERROR = traceback.format_exc()

# openpyxl rejects XML 1.0 illegal control chars (vertical tab, form feed, etc.).
# CFDI fields and OCR output occasionally smuggle one in (PDF text extraction
# is notorious for U+000B between syllables). Strip them recursively from any
# string value in the incoming payload before it reaches ws.cell.
_ILLEGAL_XLSX_CHARS = re.compile(r'[\x00-\x08\x0b-\x0c\x0e-\x1f]')

def sanitize(v):
    if isinstance(v, str):
        return _ILLEGAL_XLSX_CHARS.sub('', v)
    if isinstance(v, dict):
        return {k: sanitize(x) for k, x in v.items()}
    if isinstance(v, list):
        return [sanitize(x) for x in v]
    return v

# SMTO Brand palette
SMTO_BLACK = '050505'
SMTO_GREEN = '59D39B'
EXCEL_GREEN = '00B050'
SMTO_GREEN_DARK = '41C784'
SMTO_GREEN_SOFT = 'E8F8F0'

# Premium neutrals
BG_PAGE = 'F5F7FA'
WHITE = 'FFFFFF'
TEXT_PRIMARY = '0F172A'
TEXT_SECONDARY = '64748B'
TEXT_MUTED = '94A3B8'
BORDER_LIGHT = 'E2E8F0'
BORDER_MEDIUM = '94A3B8'
ROW_ALT = 'F8FAFC'
HEADER_BG = 'F1F5F9'

# Badge colors (soft transparent feel)
BADGE_BLUE_BG = 'EFF6FF'
BADGE_BLUE_FG = '1E40AF'
BADGE_GREEN_BG = SMTO_GREEN_SOFT
BADGE_GREEN_FG = '047857'
BADGE_GRAY_BG = 'F1F5F9'
BADGE_GRAY_FG = '475569'

# Currency-symbol lookup. Multi-divisa SIN columna extra (diseño aprobado por
# Victor Aceves 2026-07-31): el layout se queda en 22 columnas y la divisa se ve
# DENTRO de la columna MONTO M.E. vía number_format con su símbolo nativo
# (¥50 / €82.50 / RM 480.00). Se usa en:
#   1) el number_format de MONTO M.E. (columna R, misma posición de siempre);
#   2) la etiqueta del sub-renglón de propina ("Propina €8.25 EUR");
#   3) las etiquetas del bloque de T/C de la fila 7.
# Los símbolos con letras llevan espacio final para no pegarse al número.
CURRENCY_SYMBOLS = {
    'MXN': '$',   'USD': '$',    'EUR': '€',    'GBP': '£',
    'JPY': '¥',   'CNY': '¥',    'CAD': 'C$',   'AUD': 'A$',
    'CHF': 'Fr ', 'MYR': 'RM ',  'SGD': 'S$',   'HKD': 'HK$',
    'TWD': 'NT$', 'KRW': '₩',    'THB': '฿',    'VND': '₫',
    'INR': '₹',   'PHP': '₱',    'IDR': 'Rp ',  'BRL': 'R$',
    'SEK': 'kr ', 'NOK': 'kr ',  'DKK': 'kr ',  'PLN': 'zł ',
    'CZK': 'Kč ', 'HUF': 'Ft ',  'TRY': '₺',    'ILS': '₪',
    'COP': '$',   'CLP': '$',    'ARS': '$',    'PEN': 'S/ ',
}

# Monedas SIN decimales: se cotizan en unidades enteras (¥12,800, no ¥12,800.00).
CURRENCY_NO_DECIMALS = {'JPY', 'KRW', 'VND', 'CLP', 'IDR', 'HUF'}

# Último recurso si el frontend/OCR mandó el símbolo o el nombre en vez del
# código ISO ("RM", "yuan", "€"). El prompt del OCR ya pide ISO; esto blinda.
CURRENCY_ALIASES = {
    'RM': 'MYR', 'RINGGIT': 'MYR', 'RMB': 'CNY', 'YUAN': 'CNY',
    'YEN': 'JPY', 'EURO': 'EUR', 'EUROS': 'EUR', 'US$': 'USD',
    'DOLAR': 'USD', 'DOLLAR': 'USD', 'LIBRA': 'GBP', 'POUND': 'GBP',
    '€': 'EUR', '£': 'GBP', '¥': 'CNY', '₩': 'KRW', '฿': 'THB',
    '₫': 'VND', '₹': 'INR', '₱': 'PHP',
}


def normaliza_moneda(raw):
    s = str(raw or '').strip().upper()
    if not s:
        return 'MXN'
    if re.fullmatch(r'[A-Z]{3}', s):
        return s
    return CURRENCY_ALIASES.get(s, 'MXN')


def currency_symbol(code):
    """Símbolo de la divisa; si no está en el mapa usa el código ISO + espacio."""
    code = (code or 'MXN').upper()
    return CURRENCY_SYMBOLS.get(code, code + ' ')


def currency_format(code):
    """number_format de Excel para montos en `code` con su símbolo nativo. Para
    MXN devuelve exactamente el formato clásico '"$"#,##0.00' (cero cambio
    visual en reportes nacionales)."""
    code = (code or 'MXN').upper()
    # Comillas para que Excel trate el símbolo como literal — sin ellas, letras
    # como "RM" o "kr" se interpretan como códigos de formato.
    sym = currency_symbol(code).replace('"', '')
    decimals = '' if code in CURRENCY_NO_DECIMALS else '.00'
    return f'"{sym}"#,##0{decimals}'
BADGE_AMBER_BG = 'FEF3C7'
BADGE_AMBER_FG = '92400E'
BADGE_PURPLE_BG = 'F3E8FF'
BADGE_PURPLE_FG = '6B21A8'

# Colaboradores especiales — replica de COLABORADORES_ESPECIALES en App.jsx.
# Estos colaboradores reciben una celda editable de Tipo de Cambio en su Excel
# para convertir filas en USD a MXN sobre la marcha.
COLABORADORES_ESPECIALES = {
    'Alejandro Olivar',
    'Victor Aceves',
    'Daniel Covarrubias',
    'Olivia Gil',
}

def crop_logo(pil_img):
    """Remove excess whitespace/transparent padding around the logo so it
    visually aligns with surrounding type, regardless of the source PNG's
    bounding box."""
    try:
        import numpy as np
        rgba = pil_img.convert('RGBA')
        data = np.array(rgba)
        # Treat any non-transparent pixel as logo content.
        mask = data[:, :, 3] > 30
        rows = np.any(mask, axis=1)
        cols = np.any(mask, axis=0)
        if rows.any() and cols.any():
            rmin, rmax = np.where(rows)[0][[0, -1]]
            cmin, cmax = np.where(cols)[0][[0, -1]]
            pad = 4
            rmin = max(0, rmin - pad); rmax = min(data.shape[0], rmax + pad)
            cmin = max(0, cmin - pad); cmax = min(data.shape[1], cmax + pad)
            return pil_img.crop((cmin, rmin, cmax, rmax))
    except Exception:
        pass
    return pil_img

def compute_date_range(gastos):
    """Min/max of fechaCobro across gastos (falls back to fechaFac when a
    row has no cobro date). Returns ('', '') when no parseable dates exist.
    Output format: MM-DD-YY (2-digit year)."""
    dates = []
    for g in gastos:
        for key in ('fechaCobro', 'fechaFac'):
            v = g.get(key, '')
            if v and len(v) >= 8:
                # Normalize to YYYY-MM-DD for sorting
                if '-' in v and len(v) == 10 and v[4] == '-':
                    dates.append(v)  # already YYYY-MM-DD
                break
    if not dates:
        return '', ''
    dates.sort()  # ascending — smallest date first
    def fmt(d):
        # YYYY-MM-DD → DD/MM/AAAA
        return f'{d[8:10]}/{d[5:7]}/{d[0:4]}'
    return fmt(dates[0]), fmt(dates[-1])

def find_logo():
    candidates = [
        os.path.join(os.path.dirname(__file__), 'logo.png'),
        os.path.join(os.path.dirname(__file__), '..', 'public', 'logo.png'),
        '/var/task/api/logo.png',
        '/var/task/public/logo.png',
    ]
    for p in candidates:
        if os.path.exists(p):
            return p
    return None

def format_date(date_str):
    """Formato de fecha SOLO para el Excel exportado: DD/MM/AAAA.
    Acepta YYYY-MM-DD (interno) o DD/MM/YYYY (ya formateado)."""
    if not date_str: return ''
    if '-' in date_str and len(date_str) == 10:
        parts = date_str.split('-')
        # YYYY-MM-DD → DD/MM/YYYY
        return f'{parts[2]}/{parts[1]}/{parts[0]}'
    if '/' in date_str:
        parts = date_str.split('/')
        # Asume DD/MM/YYYY ya correcto; sólo normaliza el año a 4 dígitos.
        yr = parts[2] if len(parts[2]) == 4 else ('20' + parts[2])
        return f'{parts[0]}/{parts[1]}/{yr}'
    return date_str

def parse_date_obj(date_str):
    """Convierte la fecha a un datetime REAL para que Excel la trate como
    fecha (celda con number_format DD/MM/AAAA: se ve como fecha, se ordena y
    se filtra como fecha — no como texto 'General'). Acepta YYYY-MM-DD
    (interno), DD/MM/YYYY o DD/MM/YY. Devuelve None si no se puede parsear,
    para que el caller caiga al string de format_date sin perder el dato."""
    if not date_str:
        return None
    s = str(date_str).strip()
    # YYYY-MM-DD (formato interno de la app)
    if '-' in s and len(s) == 10 and s[4] == '-':
        try:
            return datetime(int(s[0:4]), int(s[5:7]), int(s[8:10]))
        except ValueError:
            return None
    # DD/MM/YYYY o DD/MM/YY (ya formateado)
    if '/' in s:
        parts = s.split('/')
        if len(parts) == 3:
            d, m, y = parts
            if len(y) == 2:
                y = '20' + y
            try:
                return datetime(int(y), int(m), int(d))
            except ValueError:
                return None
    return None

FORMA_PAGO_MAP = {
    '01': '01 - Efectivo',
    '02': '02 - Efectivo',
    '03': '03 - Transferencia',
    '04': '04 - Tarjeta de Crédito',
    '99': '99 - Por Definir',
}

def get_tipo_badge_colors(tipo):
    """Return (bg, fg) for the tipo badge.

    Paleta acordada (archivo 'Colores para tipos.xlsx'):
      VERDE    → nómina, impuestos y costos fijos recurrentes
      AZUL     → operación, viáticos y servicios (default)
      MORADO   → mantenimiento, herramientas y tecnología
      AMARILLO → papelería, uniformes, banca, préstamos y no comprobado
      GRIS     → traspasos / rechazadas (neutro)
    """
    t = (tipo or '').lower().strip()

    # ── MORADO: mantenimiento, herramientas, tecnología ──
    if any(k in t for k in ['manto', 'herramienta', 'it & sw', 'it&sw']):
        return BADGE_PURPLE_BG, BADGE_PURPLE_FG

    # ── AMARILLO: papelería, uniformes, banca, préstamos, no comprobado ──
    if any(k in t for k in ['papelería', 'papeleria', 'uniforme', 'comisión banco',
                            'comision banco', 'préstamo', 'prestamo', 'crédito',
                            'credito', 'no comprobado']):
        return BADGE_AMBER_BG, BADGE_AMBER_FG

    # ── GRIS: neutros (traspasos, rechazadas) ──
    if any(k in t for k in ['rechazada', 'devolución', 'devolucion', 'traspaso']):
        return BADGE_GRAY_BG, BADGE_GRAY_FG

    # ── VERDE: nómina, impuestos y costos fijos ──
    if any(k in t for k in ['nómina', 'nomina', 'fondo de ahorro', 'regalía', 'regalia',
                            'seguro', 'celular', 'renta oficina', 'imss', 'infonavit',
                            'isr', 'iva', 'isn']):
        return BADGE_GREEN_BG, BADGE_GREEN_FG

    # ── AZUL: operación, viáticos y servicios (default) ──
    return BADGE_BLUE_BG, BADGE_BLUE_FG

def s_border(bottom=None, top=None, left=None, right=None):
    sides = {}
    if bottom: sides['bottom'] = Side(style=bottom[0], color=bottom[1])
    if top: sides['top'] = Side(style=top[0], color=top[1])
    if left: sides['left'] = Side(style=left[0], color=left[1])
    if right: sides['right'] = Side(style=right[0], color=right[1])
    return Border(**sides)

def fill_row_bg(ws, row, start_col, end_col, color):
    for c in range(start_col, end_col + 1):
        ws.cell(row=row, column=c).fill = PatternFill('solid', start_color=color)

def style_data_cell(cell, style_type, tipo='', diff_num=None, moneda='MXN'):
    """Aplica number_format + font (y fill en badges) a una celda de datos.
    Lo comparten el renglón principal y la sub-fila de propina para que ambos
    se vean idénticos. `moneda` solo lo usa 'currency_ext' (columna MONTO
    M.E.): formatea el número con el símbolo nativo de la divisa del renglón
    (¥50 / €82.50 / RM 480.00); para MXN es el mismo formato '$' de siempre."""
    if style_type == 'currency':
        cell.number_format = '"$"#,##0.00'
        cell.font = Font(name='Calibri', size=10, color=TEXT_PRIMARY)
    elif style_type == 'currency_ext':
        cell.number_format = currency_format(moneda)
        cell.font = Font(name='Calibri', size=10, color=TEXT_PRIMARY)
    elif style_type == 'currency_bold':
        cell.number_format = '"$"#,##0.00'
        cell.font = Font(name='Calibri', size=10, bold=True, color=SMTO_BLACK)
    elif style_type == 'rfc':
        cell.font = Font(name='Calibri', size=10, bold=True, color=TEXT_PRIMARY)
    elif style_type == 'normal_bold':
        cell.font = Font(name='Calibri', size=10, bold=True, color=TEXT_PRIMARY)
    elif style_type == 'badge_tipo':
        bg_b, fg_b = get_tipo_badge_colors(tipo)
        cell.fill = PatternFill('solid', start_color=bg_b)
        cell.font = Font(name='Calibri', size=9, bold=True, color=fg_b)
    elif style_type == 'badge_pago':
        cell.fill = PatternFill('solid', start_color=BADGE_GRAY_BG)
        cell.font = Font(name='Calibri', size=9, bold=True, color=BADGE_GRAY_FG)
    elif style_type == 'tipocambio':
        cell.number_format = '#,##0.00'
        cell.font = Font(name='Calibri', size=10, color=TEXT_PRIMARY)
    elif style_type == 'date':
        cell.number_format = 'DD/MM/YYYY'
        cell.font = Font(name='Calibri', size=10, color=TEXT_PRIMARY)
    elif style_type == 'diff':
        cell.number_format = '"$"#,##0.00'
        if isinstance(diff_num, (int, float)):
            if diff_num < 0:
                cell.font = Font(name='Calibri', size=10, bold=True, color='B91C1C')
            elif diff_num > 0:
                cell.font = Font(name='Calibri', size=10, bold=True, color='15803D')
            else:
                cell.font = Font(name='Calibri', size=10, color=TEXT_PRIMARY)
        else:
            cell.font = Font(name='Calibri', size=10, color=TEXT_PRIMARY)
    else:  # 'normal'
        cell.font = Font(name='Calibri', size=10, color=TEXT_PRIMARY)

def build_workbook(gastos, colaborador='', poliza_numero='N/A', polizas_map=None):
    # Invierte POLIZAS_CLARA (nombre → folio) a (folio → nombre) para que la
    # columna USUARIO pueda resolver cada póliza al colaborador dueño.
    polizas_map = polizas_map or {}
    usuario_por_poliza = {}
    for nombre, folio in polizas_map.items():
        if folio:
            usuario_por_poliza[str(folio).strip()] = nombre
    wb = Workbook()
    # Protección a nivel libro EXPLÍCITAMENTE deshabilitada. Sin esto, algunos
    # usuarios de Windows ven el diálogo "es un archivo de solo lectura" al
    # intentar guardar — Excel interpreta cualquier elemento <fileSharing> o
    # WorkbookProtection con flags ambiguos como "read-only recommended".
    wb.security = WorkbookProtection(
        lockStructure=False,
        lockWindows=False,
        lockRevision=False,
        workbookPassword=None,
        revisionsPassword=None,
    )
    # Forzar recálculo de TODAS las fórmulas al abrir (KPIs, TOTAL, SUM). openpyxl
    # no precalcula valores; sin esto, un visor que respeta el caché muestra 0.
    wb.calculation.fullCalcOnLoad = True
    ws = wb.active
    ws.title = 'Reporte SMTO'
    ws.sheet_view.showGridLines = False
    ws.protection.sheet = False  # make the sheet fully editable after export
    ws.protection.enabled = False  # belt-and-suspenders: also drop the element

    # Column widths — semantic (wide CONCEPTO + supplier, narrow dates).
    # Layout: A spacer, B RFC, C PROVEEDOR, D TIPO, E PÓLIZA, F FACTURA,
    # G F.FACTURA, H F.COBRO, I CONCEPTO, J IMPORTE, K IVA, L ISR (nueva),
    # M ISH/IEPS, N RETENCIÓN, O TOTAL, P FORMA PAGO, Q BANCO, R MONTO USD,
    # S T/C, T USUARIO, U COBRADO, V FACTURADO, W DIFERENCIA, X spacer.
    col_widths = {
        'A': 3, 'B': 15, 'C': 30, 'D': 11, 'E': 10, 'F': 14, 'G': 11, 'H': 11,
        'I': 28, 'J': 12, 'K': 11, 'L': 11, 'M': 11, 'N': 11, 'O': 18,
        'P': 15, 'Q': 22, 'R': 13, 'S': 13, 'T': 24,
        'U': 14, 'V': 22, 'W': 14, 'X': 3, 'Y': 16,
    }
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # Paint full background light gray — scales with row count so the
    # outer spacer cols past the totals/footer still inherit BG_PAGE.
    nrows_painted = max(80, 40 + len(gastos))
    for r in range(1, nrows_painted):
        fill_row_bg(ws, r, 1, 24, BG_PAGE)

    # ═══ HEADER (rows 1-2) — title + colaborador labels + fields ═══
    ws.row_dimensions[1].height = 50
    ws.row_dimensions[2].height = 36

    # Title — centered across D1:G2; pairs with the 64px logo at B1
    ws.merge_cells('D1:G2')
    title = ws['D1']
    title.value = 'Reporte de Gastos'
    title.font = Font(name='Aptos', size=26, bold=True, color=SMTO_BLACK)
    title.alignment = Alignment(horizontal='center', vertical='center')
    title.fill = PatternFill('solid', start_color=BG_PAGE)

    # Right-side form labels — merged H:I, right-aligned. Sin el merge, el texto
    # largo ("Nombre de colaborador:") se recortaba contra el título (D1:G2) a
    # la izquierda. Con H:I (ancho 11+28) cabe completo y termina justo antes
    # del campo de valor en J.
    ws.merge_cells('H1:I1')
    lbl1 = ws['H1']
    lbl1.value = 'Nombre de colaborador:'
    lbl1.font = Font(name='Aptos', size=11, bold=True, color=TEXT_SECONDARY)
    lbl1.alignment = Alignment(horizontal='right', vertical='center')
    lbl1.fill = PatternFill('solid', start_color=BG_PAGE)
    ws['I1'].fill = PatternFill('solid', start_color=BG_PAGE)

    ws.merge_cells('H2:I2')
    lbl2 = ws['H2']
    lbl2.value = 'Fecha de viaje:'
    lbl2.font = Font(name='Aptos', size=11, bold=True, color=TEXT_SECONDARY)
    lbl2.alignment = Alignment(horizontal='right', vertical='center')
    lbl2.fill = PatternFill('solid', start_color=BG_PAGE)
    ws['I2'].fill = PatternFill('solid', start_color=BG_PAGE)

    # Right-side form fields (cols J:M) — compact 11pt
    ws.merge_cells('J1:M1')
    f1 = ws['J1']
    f1.value = colaborador  # filled programmatically from the selected colaborador
    f1.font = Font(name='Aptos', size=11, color=TEXT_PRIMARY)
    f1.fill = PatternFill('solid', start_color=WHITE)
    f1.border = Border(bottom=Side(style='thin', color=BORDER_LIGHT))
    f1.alignment = Alignment(horizontal='left', vertical='center', indent=1)

    ws.merge_cells('J2:M2')
    f2 = ws['J2']
    fecha_min, fecha_max = compute_date_range(gastos)
    if fecha_min:
        f2.value = f'DE: {fecha_min}  A: {fecha_max}'
    else:
        f2.value = ''
    f2.font = Font(name='Aptos', size=11, color=TEXT_PRIMARY)
    f2.fill = PatternFill('solid', start_color=WHITE)
    f2.alignment = Alignment(horizontal='left', vertical='center', indent=1)

    # Row 3 spacer + Row 4 closing line under the header section.
    # H3 carries a thin EXCEL_GREEN underline below the "Nombre/Fecha de viaje"
    # labels; row 4 itself is a 1pt-tall band sandwiched between a thin top
    # green border and a medium bottom green border so it reads as one bar.
    ws.row_dimensions[3].height = 10
    ws.row_dimensions[4].height = 1
    ws['H3'].border = Border(bottom=Side(style='thin', color=EXCEL_GREEN))
    for c in range(2, 24):
        cell = ws.cell(row=4, column=c)
        cell.fill = PatternFill('solid', start_color=BG_PAGE)
        cell.border = Border(
            top=Side(style='thin', color=EXCEL_GREEN),
            bottom=Side(style='medium', color=EXCEL_GREEN),
        )

    # ═══ KPI CARDS (rows 5-6) ═══
    ws.row_dimensions[5].height = 32  # KPI VALUES
    ws.row_dimensions[6].height = 16  # KPI LABELS

    # Pre-compute the data range so the KPI cards AND the bottom totals row
    # share the same =SUM() references. Counts every main row + any propina
    # sub-row, clamps to data_first when no gastos so the SUM range stays
    # valid (I10:I10 → 0).
    data_first = 10
    has_propina = lambda g: (g.get('montoPropina', 0) or 0) > 0 or (g.get('propinaExtranjero', 0) or 0) > 0
    data_rows_total = len(gastos) + sum(1 for g in gastos if has_propina(g))
    data_last = max(data_first, data_first + data_rows_total - 1)
    num_facturas = len(gastos)

    # (col_start, col_end, label, value, number_format, value_color)
    # KPI cards are live Excel formulas referencing the data band so the
    # numbers always match the bottom totals row. REGISTROS stays a static count.
    # 7 tarjetas sobre B:W (22 cols): la headline TOTAL FACTURADO ocupa 4 cols
    # y las demás 3 c/u (4 + 6×3 = 22), así DIFERENCIA es del mismo tamaño que el
    # resto (ya no es un rectángulo grande) y la banda queda simétrica.
    # Columnas de datos tras ISR/ISH: TOTAL=O, IVA=K, ISR=L, RETENCIÓN=N,
    # MONTO USD=R, DIFERENCIA=W, CONCEPTO=I.
    #  - TOTAL FACTURADO = total SIN propinas (resta las sub-filas "Propina*").
    #  - TOTAL COBRADO   = total CON propinas (todo lo que cargó la tarjeta).
    # Divisas extranjeras presentes en el reporte (orden estable: USD primero
    # por ser la más común, resto alfabético). Con UNA divisa el reporte se ve
    # EXACTAMENTE como el clásico de USD; con varias cambian solo la tarjeta
    # R-T y los formatos de la columna MONTO M.E. — nunca el layout.
    monedas_ext = sorted({
        normaliza_moneda(g.get('monedaCodigo') or g.get('moneda'))
        for g in gastos
    } - {'MXN', 'XXX'}, key=lambda m: (m != 'USD', m))

    suma_total   = f'SUM(O{data_first}:O{data_last})'
    suma_propina = f'SUMIF(I{data_first}:I{data_last},"Propina*",O{data_first}:O{data_last})'

    # Tarjeta R-T ("USD" en el diseño clásico):
    #  - 0 o 1 divisa → misma SUM(R) de siempre, con la etiqueta y el símbolo
    #    de ESA divisa (un viaje a Malasia ve "MYR / RM 480.00").
    #  - 2+ divisas → SUM(R) sumaría peras con manzanas (100 EUR + 400 MYR =
    #    "500"); en su lugar suma el equivalente EN PESOS de los renglones
    #    extranjeros: las filas con T/C (col S > 0) tienen su TOTAL (col O) ya
    #    convertido a MXN por la fórmula nativa×T/C.
    if len(monedas_ext) > 1:
        kpi_fx_lbl = 'M. EXTRANJERA (MXN)'
        kpi_fx_val = f'=SUMIF(S{data_first}:S{data_last},">0",O{data_first}:O{data_last})'
        kpi_fx_fmt = '"$"#,##0.00'
    else:
        kpi_fx_lbl = monedas_ext[0] if monedas_ext else 'USD'
        kpi_fx_val = f'=SUM(R{data_first}:R{data_last})'
        kpi_fx_fmt = currency_format(kpi_fx_lbl)

    kpis = [
        ('B', 'E', 'TOTAL FACTURADO', f'={suma_total}-{suma_propina}', '"$"#,##0.00', SMTO_GREEN),
        ('F', 'H', 'TOTAL COBRADO',   f'={suma_total}',                '"$"#,##0.00', SMTO_GREEN),
        ('I', 'K', 'IVA TOTAL',       f'=SUM(K{data_first}:K{data_last})', '"$"#,##0.00', SMTO_BLACK),
        # RETENCIONES = ISR retenido (col L) + RETENCIÓN (col N): el ISR tiene su
        # propia columna pero sigue siendo una retención.
        ('L', 'N', 'RETENCIONES',     f'=SUM(L{data_first}:L{data_last})+SUM(N{data_first}:N{data_last})', '"$"#,##0.00', SMTO_BLACK),
        ('O', 'Q', 'REGISTROS',       num_facturas,                        '0',           SMTO_BLACK),
        ('R', 'T', kpi_fx_lbl,        kpi_fx_val,                          kpi_fx_fmt,    SMTO_GREEN),
        # DIFERENCIA cobrado vs facturado — SOLO renglones Clara MXN Credito.
        ('U', 'W', 'DIFERENCIA',      f'=SUM(W{data_first}:W{data_last})', '"$"#,##0.00', SMTO_GREEN),
    ]

    for col_start, col_end, label, value, fmt, value_color in kpis:
        # Value row (row 5) — anchor cell carries the value/font/fill; the
        # right-edge border has to also be applied to the LAST cell of the
        # merge because openpyxl renders borders per cell, not per merge.
        ws.merge_cells(f'{col_start}5:{col_end}5')
        v_cell = ws[f'{col_start}5']
        v_cell.value = value
        v_cell.number_format = fmt
        v_cell.font = Font(name='Aptos', size=20, bold=True, color=value_color)
        v_cell.alignment = Alignment(horizontal='left', vertical='center', indent=2)
        v_cell.fill = PatternFill('solid', start_color=WHITE)
        v_cell.border = Border(
            left=Side(style='medium', color=EXCEL_GREEN),
            top=Side(style='medium', color=EXCEL_GREEN),
            right=Side(style='medium', color=EXCEL_GREEN),
            bottom=Side(style='thin', color=BORDER_LIGHT),
        )
        ws[f'{col_end}5'].border = Border(
            right=Side(style='medium', color=EXCEL_GREEN),
            top=Side(style='medium', color=EXCEL_GREEN),
            bottom=Side(style='thin', color=BORDER_LIGHT),
        )

        # Label row (row 6) — same trick: border the merge end-cell too.
        ws.merge_cells(f'{col_start}6:{col_end}6')
        l_cell = ws[f'{col_start}6']
        l_cell.value = label
        l_cell.font = Font(name='Aptos', size=9, bold=True, color=TEXT_MUTED)
        l_cell.alignment = Alignment(horizontal='left', vertical='center', indent=2)
        l_cell.fill = PatternFill('solid', start_color=WHITE)
        l_cell.border = Border(
            left=Side(style='medium', color=EXCEL_GREEN),
            top=Side(style='thin', color=BORDER_LIGHT),
            right=Side(style='medium', color=EXCEL_GREEN),
            bottom=Side(style='medium', color=EXCEL_GREEN),
        )
        ws[f'{col_end}6'].border = Border(
            right=Side(style='medium', color=EXCEL_GREEN),
            top=Side(style='thin', color=BORDER_LIGHT),
            bottom=Side(style='medium', color=EXCEL_GREEN),
        )

    # Middle cells of each merged KPI range need explicit borders too —
    # without them, some Excel/LibreOffice versions render visible gaps
    # inside the cards even though the anchor + end cells have borders.
    MED_GREEN  = Side(style='medium', color=EXCEL_GREEN)
    THIN_LIGHT = Side(style='thin',   color=BORDER_LIGHT)
    # Middle cells por tarjeta: B-E→C,D · F-H→G · I-K→J · L-N→M · O-Q→P ·
    # R-T→S · U-W→V
    for col_letter in ('C', 'D', 'G', 'J', 'M', 'P', 'S', 'V'):
        ws[f'{col_letter}5'].border = Border(
            left=MED_GREEN, right=MED_GREEN, top=MED_GREEN, bottom=THIN_LIGHT,
        )
        ws[f'{col_letter}6'].border = Border(
            left=MED_GREEN, right=MED_GREEN, top=THIN_LIGHT, bottom=MED_GREEN,
        )

    # ═══ TIPO DE CAMBIO EDITABLE (solo colaboradores especiales) ═══
    # UNA celda editable POR CADA divisa del reporte, en la fila 7 (fuera de la
    # tabla — el layout de 22 columnas NO se toca). Cada renglón se convierte
    # con la celda de SU divisa: los euros ya no usan la tasa del dólar.
    #
    # Compatibilidad: el primer par (USD, o la única divisa) se ancla EXACTO en
    # la posición clásica — etiqueta O7:Q7 + input R7 — así el caso común (una
    # sola divisa) es indistinguible del formato v8.92 aprobado. Los pares
    # extra crecen a la IZQUIERDA (N7←, J7←, F7←); con 5+ divisas se compactan
    # a etiqueta+input de 1 columna c/u.
    is_especial = (colaborador or '').strip() in COLABORADORES_ESPECIALES
    tc_refs = {}   # 'EUR' -> '$N$7'

    if is_especial:
        ws.row_dimensions[7].height = 30
        amber_side = Side(style='medium', color='F59E0B')

        # Semilla por divisa: la tasa que ya traiga algún gasto (la conciliación
        # la deriva del cargo real: MXN ÷ monto nativo). USD conserva su default
        # histórico de 17.50; para otras divisas sin dato la celda queda VACÍA
        # a propósito — inventar una tasa daría un reporte plausible pero
        # incorrecto, y el hueco grita "captúrame".
        lista_tc = monedas_ext or ['USD']  # sin divisas: celda clásica de USD
        seed_tc = {}
        for g in gastos:
            m = normaliza_moneda(g.get('monedaCodigo') or g.get('moneda'))
            rate = g.get('tipoCambio') or 0
            if m in lista_tc and m not in seed_tc and rate > 0:
                seed_tc[m] = round(float(rate), 4)
        if 'USD' in lista_tc:
            seed_tc.setdefault('USD', 17.50)

        ancho = 4 if len(lista_tc) <= 4 else 2   # cols por par (etiqueta+input)
        n_pares = min(len(lista_tc), 4 if ancho == 4 else 8)
        for i, code in enumerate(lista_tc[:n_pares]):
            col_in  = 18 - ancho * i             # input: R7, luego N7/J7/F7…
            col_lbl = col_in - (ancho - 1)

            if ancho == 4:
                ws.merge_cells(start_row=7, start_column=col_lbl,
                               end_row=7, end_column=col_in - 1)
                texto = ('TIPO DE CAMBIO USD →' if len(lista_tc) == 1 and code == 'USD'
                         else f'T/C {code} →')
            else:
                texto = f'{code}→'
            lbl_tc = ws.cell(row=7, column=col_lbl, value=texto)
            lbl_tc.font = Font(name='Aptos', size=11 if ancho == 4 else 9,
                               bold=True, color=BADGE_AMBER_FG)
            lbl_tc.alignment = Alignment(horizontal='right', vertical='center', indent=1)
            # openpyxl no propaga el fill dentro de un merge: pintar cada celda.
            for c in range(col_lbl, col_in):
                ws.cell(row=7, column=c).fill = PatternFill('solid', start_color=BADGE_AMBER_BG)

            tc_input = ws.cell(row=7, column=col_in)
            if code in seed_tc:
                tc_input.value = seed_tc[code]
            tc_input.number_format = '#,##0.0000'
            tc_input.font = Font(name='Aptos', size=16 if ancho == 4 else 12,
                                 bold=True, color=BADGE_AMBER_FG)
            tc_input.alignment = Alignment(horizontal='center', vertical='center')
            tc_input.fill = PatternFill('solid', start_color='FDE68A')  # amber-200
            tc_input.border = Border(
                left=amber_side, right=amber_side, top=amber_side, bottom=amber_side,
            )
            tc_refs[code] = f'${get_column_letter(col_in)}$7'

    # Row 7 (cuando NO es especial) small gap + Row 8 pre-table spacer
    if not is_especial:
        ws.row_dimensions[7].height = 8
    ws.row_dimensions[8].height = 14

    # ═══ TABLE HEADER (row 9) — green text, mostly centered ═══
    # 34px cabe las 2 líneas del header FACTURADO "(incluye propinas)".
    ws.row_dimensions[9].height = 34

    # FACTURADO lleva un sub-rótulo "(incluye propinas)" en segunda línea para
    # avisar que su total suma las propinas (así cuadra contra COBRADO).
    # 'MONTO M.E.' (antes 'MONTO USD') — MISMA columna R, solo cambia el rótulo:
    # el monto va en la divisa original del ticket con su símbolo en el formato
    # de la celda (¥50 / €82.50 / RM 480.00). El import de la app acepta ambos
    # encabezados.
    headers = ['RFC', 'PROVEEDOR', 'TIPO', 'PÓLIZA', 'FACTURA', 'F. FACTURA', 'F. COBRO', 'CONCEPTO', 'IMPORTE', 'IVA', 'ISR', 'ISH/IEPS', 'RETENCIÓN', 'TOTAL', 'FORMA PAGO', 'BANCO', 'MONTO M.E.', 'T/C', 'USUARIO', 'COBRADO', 'FACTURADO\n(incluye propinas)', 'DIFERENCIA']
    # PROVEEDOR and CONCEPTO stay left-aligned; the rest center.
    left_align_headers = {'PROVEEDOR', 'CONCEPTO'}

    for i, h in enumerate(headers):
        col = i + 2
        cell = ws.cell(row=9, column=col, value=h)
        cell.font = Font(name='Aptos', size=11, bold=True, color=EXCEL_GREEN)
        is_left = h in left_align_headers
        cell.alignment = Alignment(
            horizontal='left' if is_left else 'center',
            vertical='center',
            indent=2 if is_left else 0,
            wrap_text='\n' in h,  # FACTURADO usa 2 líneas
        )
        cell.fill = PatternFill(fill_type=None)
        # Top + bottom medium green on every header; left edge on B9 (first),
        # right edge on W9 (last, col 23) so the band reads as one bordered strip.
        cell.border = Border(
            top=Side(style='medium', color=EXCEL_GREEN),
            bottom=Side(style='medium', color=EXCEL_GREEN),
            left=Side(style='medium', color=EXCEL_GREEN) if col == 2 else None,
            right=Side(style='medium', color=EXCEL_GREEN) if col == 23 else None,
        )

    # Autofiltro sobre el encabezado (fila 9) + todas las filas de datos
    # (B9:U{data_last}), para que el usuario filtre cualquier columna —
    # CONCEPTO, TIPO, PROVEEDOR, fechas, etc.— con los dropdowns nativos.
    # data_last se calculó arriba (cubre filas principales + sub-filas de
    # propina) y NO incluye la banda de TOTAL CUENTA, que queda fuera del
    # filtro. Nota: TOTAL usa SUM (no SUBTOTAL), así que la fila de totales
    # no cambia al filtrar — es el comportamiento previo, intacto.
    ws.auto_filter.ref = f'B9:W{data_last}'

    # ¿Hubo conciliación bancaria en esta sesión? Si CUALQUIER fila trae un
    # montoCobrado real del CSV (o hizoMatch del frontend), el reporte está
    # en "modo conciliado": las filas Clara que NO encontraron su cargo ya no
    # fabrican COBRADO = factura+propina — eso pintaba DIFERENCIA $0.00 y
    # escondía faltantes reales (hotel facturado $7,344 vs cobrado $7,334
    # salía como si cuadrara). Ahora muestran "SIN CONCILIAR" en COBRADO.
    # Sin corrida de banco (export temprano), se conserva el estimado
    # histórico para no llenar el reporte de marcas.
    algo_conciliado = any(
        (g.get('montoCobrado') or 0) > 0 or g.get('hizoMatch')
        for g in gastos
    )

    # ═══ DATA ROWS (row 10+) ═══
    # Acumuladores para la banda de totales de las 3 columnas de conciliación.
    # Se calculan en Python (no con fórmulas SUM) por la MISMA razón que las
    # celdas por-renglón COBRADO/FACTURADO/DIFERENCIA se escriben como VALOR:
    # así la fila de totales muestra el número correcto en CUALQUIER visor
    # (Excel, Numbers, Quick Look, Google Sheets) sin depender del recálculo.
    #  - total_cobrado_band   = Σ COBRADO (cargo real de la tarjeta, ya incluye tip)
    #  - total_facturado_band = Σ (FACTURADO + propina) → para que cuadre vs COBRADO
    #  - DIFERENCIA total      = total_cobrado_band − total_facturado_band
    total_cobrado_band = 0.0
    total_facturado_band = 0.0
    row = 10
    for idx, g in enumerate(gastos):
        ws.row_dimensions[row].height = 15
        is_alt = (idx % 2 == 1)
        bg = ROW_ALT if is_alt else WHITE
        # XML corrupto rescatado por OCR: la fila NO tiene su CFDI XML. Se tinta
        # de ámbar muy claro y se marca "⚠ Falta XML" a la derecha de la tabla
        # (columna Y) para que el revisor sepa que falta descargar/adjuntar el XML.
        xml_faltante = bool(g.get('xmlFaltante'))
        if xml_faltante:
            bg = 'FEF3C7'  # ámbar claro

        # Fechas como datetime REAL (Excel las trata como fecha, no como texto
        # 'General'): se ordenan y filtran como fecha. Si no parsean, cae al
        # string DD/MM/AAAA de format_date para no perder el dato.
        fecha_fac_obj   = parse_date_obj(g.get('fechaFac', ''))
        fecha_cobro_obj = parse_date_obj(g.get('fechaCobro', ''))
        fecha_fac   = fecha_fac_obj   if fecha_fac_obj   else format_date(g.get('fechaFac', ''))
        fecha_cobro = fecha_cobro_obj if fecha_cobro_obj else format_date(g.get('fechaCobro', ''))
        forma = FORMA_PAGO_MAP.get(g.get('formaPago', '04'), g.get('formaPago', ''))
        importe_raw = round(g.get('importe', 0), 2)
        iva = round(g.get('iva', 0), 2)
        # ISR e ISH/IEPS en columnas SEPARADAS, con SIGNOS distintos:
        #  - isr_amt = ISR RETENIDO (retencionISR) → RESTA del total (es una
        #    retención: p.ej. Volare 494.12+79.06−6.18 = 567.00).
        #  - ish     = ISH (locales de hoteles) + IEPS (combustible) → SUMAN.
        # La columna RETENCIÓN ya NO incluye el ISR (tiene su propia columna),
        # así que = retenciones totales − ISR retenido (para no contarlo doble).
        # `ishIeps` cae a isrTrasladado para filas legadas sin desglose.
        isr_amt = round(g.get('retencionISR', 0) or 0, 2)
        ish = round(g.get('ishIeps', g.get('isrTrasladado', 0)) or 0, 2)
        ret = round((g.get('retenciones', 0) or 0) - (g.get('retencionISR', 0) or 0), 2)
        total = round(g.get('totalCFDI', 0) + g.get('montoPropina', 0), 2)
        tipo = g.get('tipo', 'Consumo')
        monto_usd_raw = round(g.get('montoUSD', 0) or 0, 2)
        tipo_cambio = round(g.get('tipoCambio', 0) or 0, 2)
        # Valores en USD preservados desde parseCFDI (importe/iva/ret en la
        # gasto vienen en 0 hasta que se aplique el T/C en el Excel). El
        # bloque is_usd_row de abajo los multiplica por la celda editable.
        importe_usd_raw = round(g.get('importeUSD', 0) or 0, 2)
        iva_usd_raw     = round(g.get('ivaUSD', 0) or 0, 2)
        ret_usd_raw     = round(g.get('retencionesUSD', 0) or 0, 2)
        # Per-row póliza (sólo se llena cuando el colaborador es especial
        # y corrió el cotejo con su Saldos); para todos los demás cae al
        # folio Clara global del colaborador.
        _praw = g.get('polizaNumero') or poliza_numero
        # PÓLIZA como TEXTO para conservar los ceros a la izquierda (p.ej.
        # "0036234032", "0610"). int() los borraba ("0036234032" → 36234032) y
        # además rompía el lookup de USUARIO contra POLIZAS_CLARA, cuyos folios
        # sí traen ceros.
        poliza_row = '' if _praw is None else str(_praw).strip()

        # Propina-row fields. When the gasto has any propina, the main row
        # shows the NET amount (total − tip) so main_row + propina_sub_row
        # always sums to the full charge in SUM(I) / SUM(L) / SUM(N).
        # Basis is totalCFDI − montoPropina (and montoExtranjero −
        # propinaExtranjero on the foreign side). Both fields are clamped
        # to ≥ 0.
        propina_mxn = round(g.get('montoPropina', 0) or 0, 2)
        propina_ext = round(g.get('propinaExtranjero', 0) or 0, 2)
        moneda_code = normaliza_moneda(g.get('monedaCodigo') or g.get('moneda'))
        total_mxn   = round(g.get('totalCFDI', 0) or 0, 2)
        monto_ext   = round(g.get('montoExtranjero', 0) or g.get('montoUSD', 0) or 0, 2)

        if propina_mxn > 0 or propina_ext > 0:
            # El renglón principal mantiene los montos REALES del CFDI:
            # IMPORTE = SubTotal (sin IVA), de modo que L = I+J-K = totalCFDI
            # tal y como aparece en la factura ($410, $842, etc).
            # La propina vive en su propio sub-renglón con su L propio.
            importe   = importe_raw
            monto_usd = monto_usd_raw
        else:
            importe   = importe_raw
            monto_usd = monto_usd_raw

        # Passthrough de divisa: si el colaborador es especial y la factura
        # viene en moneda extranjera, IMPORTE / IVA / RETENCIÓN / T/C usan
        # formulas que multiplican el valor nativo por la celda editable DE SU
        # DIVISA (los euros con la celda de EUR, nunca con la del dólar).
        # Cambiar esa celda recalcula al instante todas las filas de esa
        # moneda. importe/iva/ret en la gasto vienen en 0 (la UI los muestra
        # así hasta tener T/C); los desgloses nativos viajan en importeUSD /
        # ivaUSD / retencionesUSD (nombres históricos: hoy significan "en la
        # divisa del ticket", no dólares).
        tc_ref = tc_refs.get(moneda_code) if moneda_code not in ('MXN', 'XXX') else None
        is_usd_row = (
            tc_ref is not None
            and (importe_usd_raw > 0 or iva_usd_raw > 0 or ret_usd_raw > 0
                 or monto_usd_raw > 0 or monto_ext > 0)
        )
        if is_usd_row:
            # Fallback: si por alguna razón no llegaron los desgloses nativos,
            # asumimos que el total nativo ES el subtotal (IVA/Ret = 0) — el
            # importe en MXN siempre coincidirá con el MONTO M.E. visible.
            usd_imp = importe_usd_raw if importe_usd_raw > 0 else (monto_usd_raw or monto_ext)
            usd_iva = iva_usd_raw
            usd_ret = ret_usd_raw
            importe_val = f'={usd_imp}*{tc_ref}'
            iva_val     = f'={usd_iva}*{tc_ref}' if usd_iva > 0 else 0
            ret_val     = f'={usd_ret}*{tc_ref}' if usd_ret > 0 else 0
            tc_val      = f'={tc_ref}'
        else:
            importe_val = importe
            iva_val     = iva
            ret_val     = ret
            tc_val      = tipo_cambio

        # Column order matches the headers. PROVEEDOR and CONCEPTO are the only
        # left-aligned cells; everything else centers per the reference.
        banco = (g.get('banco') or '').strip()
        # Fallback de banco: si el gasto NO tiene banco asignado (típico de
        # data legacy creada antes de v7.81 o cuando el colaborador no es
        # especial y nunca cotejó con Saldos) y el colaborador NO es especial,
        # defaultea a Clara MXN Credito (misma lógica que el frontend en
        # defaultBancoFor). Sin esto, la reconciliación cobrado vs facturado
        # quedaba vacía aunque el colaborador usa Clara por default.
        if not banco and colaborador and colaborador not in COLABORADORES_ESPECIALES:
            banco = 'Clara MXN Credito'
        # USUARIO se resuelve por el folio (poliza_row) usando el mapa
        # invertido de POLIZAS_CLARA. Si el folio del Saldos no matchea
        # ningún colaborador del mapa, cae al nombre del colaborador dueño
        # del reporte (e.g., 'Alejandro Olivar' para sus propias facturas).
        usuario = usuario_por_poliza.get(str(poliza_row).strip()) or colaborador or ''
        # Reconciliación cobrado vs facturado — solo aplica para renglones
        # cuya tarjeta es Clara MXN Credito. Para los demás (BBVA, Monex,
        # Kapital, USD), las 3 celdas quedan vacías (no se reconcilian con
        # Clara CSV en este flujo).
        #
        # IMPORTANTE: FACTURADO muestra el monto de la factura SIN propina
        # (el CFDI no incluye el tip). La propina se contempla en la fórmula
        # de DIFERENCIA: cobrado − (facturado + propina). Así un tip legítimo
        # nunca aparece como discrepancia, y sólo se marca en rojo cuando el
        # banco cargó más que factura + propina.
        es_clara = banco.lower() == 'clara mxn credito'
        if es_clara:
            propina_actual = round(g.get('montoPropina', 0) or 0, 2)
            # FACTURADO = monto de la factura SIN propina (el CFDI no incluye el tip).
            facturado = round(g.get('montoFacturado', 0) or g.get('totalCFDI', 0) or 0, 2)
            # COBRADO = monto REAL que cargó la tarjeta (del CSV de Clara, guardado
            # al validar banco). Es INDEPENDIENTE de factura+propina, por eso
            # DIFERENCIA puede detectar un excedente real. Si la fila no se validó
            # contra el banco (montoCobrado=0), cae al estimado factura+propina,
            # que cuadra → DIFERENCIA 0 hasta que se concilie con el banco.
            monto_cobrado = round(g.get('montoCobrado', 0) or 0, 2)
            if monto_cobrado > 0:
                # Fila conciliada: COBRADO es el cargo REAL de la tarjeta.
                # DIFERENCIA = cobrado − (facturado + propina). Se escribe el
                # VALOR calculado (NO una fórmula viva) para que la celda
                # muestre el número correcto en CUALQUIER visor — Excel,
                # Numbers, Quick Look, Google Sheets — sin depender de que
                # recalcule fórmulas. Resta la propina sólo cuando existe.
                cobrado = monto_cobrado
                diff_num = round(cobrado - facturado - propina_actual, 2)
                diferencia = diff_num
            elif algo_conciliado:
                # Hubo corrida de banco y ESTA fila no encontró su cargo:
                # decirlo explícitamente en vez de fingir que cuadra en $0.
                cobrado = 'SIN CONCILIAR'
                diff_num = None
                diferencia = ''
            else:
                # Export sin validación bancaria: estimado factura+propina
                # (comportamiento histórico — DIFERENCIA 0 hasta conciliar).
                cobrado = round((g.get('totalCFDI', 0) or 0) + propina_actual, 2)
                diff_num = round(cobrado - facturado - propina_actual, 2)
                diferencia = diff_num
            # Acumular para la banda de totales. FACTURADO suma la propina para
            # que cuadre contra COBRADO (que ya la trae en el cargo de tarjeta).
            # Los renglones "SIN CONCILIAR" suman su FACTURADO+propina pero NO su
            # COBRADO (no hay cargo real) → la DIFERENCIA total delata el faltante.
            total_facturado_band += facturado + propina_actual
            if isinstance(cobrado, (int, float)):
                total_cobrado_band += cobrado
        else:
            facturado = ''
            cobrado = ''
            diff_num = None
            diferencia = ''
        # RFC de facturas extranjeras (sin RFC mexicano): se muestra "N/A"
        # centrado. Los RFC reales van alineados a la izquierda como siempre.
        rfc_raw = g.get('rfc', '')
        rfc_is_na = str(rfc_raw).strip().upper() in ('NA', 'N/A')
        rfc_disp = 'N/A' if rfc_is_na else rfc_raw
        rfc_align = 'center' if rfc_is_na else 'left'
        # Layout tras agregar ISR (L=12) e ISH/IEPS (M=13): RETENCIÓN→N(14),
        # TOTAL→O(15), FORMA→P(16), BANCO→Q(17), MONTO USD→R(18), T/C→S(19),
        # USUARIO→T(20), COBRADO→U(21), FACTURADO→V(22), DIFERENCIA→W(23).
        cells = [
            (2,  rfc_disp,               rfc_align, 'rfc'),
            # PROVEEDOR: SIEMPRE en MAYÚSCULAS en el Excel, sin importar la fuente
            # (XML emisor, OCR, manual) ni el usuario de la plantilla.
            (3,  str(g.get('proveedor', '') or '').upper(), 'left', 'normal_bold'),
            (4,  tipo,                   'center', 'badge_tipo'),
            (5,  poliza_row,             'center', 'normal'),
            (6,  g.get('noFactura', ''), 'center', 'normal'),
            (7,  fecha_fac,              'center', 'date'),
            (8,  fecha_cobro,            'center', 'date'),
            (9,  g.get('concepto', ''),  'left',   'normal'),
            (10, importe_val,            'center', 'currency'),
            (11, iva_val,                'center', 'currency'),
            (12, isr_amt,                'center', 'currency'),   # ISR retenido (RESTA)
            (13, ish,                    'center', 'currency'),   # ISH/IEPS (SUMA)
            (14, ret_val,                'center', 'currency'),
            # TOTAL = IMPORTE + IVA − ISR + ISH/IEPS − RETENCIÓN (fórmula viva).
            # El ISR (L) y la RETENCIÓN (N) RESTAN; ISH/IEPS (M) SUMA. Así
            # hoteles/combustible/honorarios y retenciones cuadran exacto con el
            # total del CFDI. Recalcula si se editan J/K/L/M/N o con fórmulas USD.
            (15, f'=J{row}+K{row}-L{row}+M{row}-N{row}', 'center', 'currency_bold'),
            (16, forma,                  'center', 'badge_pago'),
            (17, banco,                  'center', 'normal'),
            # MONTO M.E.: monto en la divisa ORIGINAL del ticket, con su
            # símbolo en el number_format ('currency_ext'). Prefiere
            # montoExtranjero y cae a montoUSD (reportes/gastos viejos).
            (18, monto_usd or monto_ext, 'center', 'currency_ext'),
            (19, tc_val,                 'center', 'tipocambio'),
            (20, usuario,                'center', 'normal'),
            (21, cobrado,                'center', 'currency'),
            (22, facturado,              'center', 'currency'),
            (23, diferencia,             'center', 'diff'),
        ]

        for col, val, align, style_type in cells:
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = PatternFill('solid', start_color=bg)
            cell.border = Border(bottom=Side(style='hair', color=BORDER_LIGHT))
            cell.alignment = Alignment(
                horizontal=align,
                vertical='center',
                indent=2 if align == 'left' else 0
            )
            style_data_cell(cell, style_type, tipo=tipo, diff_num=diff_num,
                            moneda=moneda_code)

        # Marca de fila no conciliada: COBRADO dice "SIN CONCILIAR" en ámbar
        # bold (mismo lenguaje visual que "⚠ Falta XML") para que el revisor
        # la distinga de un $0.00 legítimo.
        if cobrado == 'SIN CONCILIAR':
            c_sin = ws.cell(row=row, column=21)
            c_sin.number_format = '@'
            c_sin.font = Font(name='Calibri', size=8.5, bold=True, color='B45309')
            c_sin.fill = PatternFill('solid', start_color='FEF3C7')

        # Side spacer cells keep page bg through the data band.
        ws.cell(row=row, column=1).fill = PatternFill('solid', start_color=BG_PAGE)
        ws.cell(row=row, column=24).fill = PatternFill('solid', start_color=BG_PAGE)

        # Marcador "Falta XML" a la derecha de la tabla (columna Y = 25). Solo
        # aparece en filas rescatadas por OCR sin su CFDI XML.
        if xml_faltante:
            mark = ws.cell(row=row, column=25, value='⚠ Falta XML')
            mark.fill = PatternFill('solid', start_color='FEF3C7')
            mark.font = Font(name='Calibri', size=9, bold=True, color='B45309')
            mark.alignment = Alignment(horizontal='left', vertical='center', indent=1)

        row += 1

        # ── Optional propina sub-row ──
        # La propina es un renglón COMPLETO que HEREDA del padre: mismo RFC,
        # mismo PROVEEDOR (no dice "Propina" en esa columna), mismo TIPO, misma
        # PÓLIZA, misma FORMA PAGO, mismo BANCO, mismo USUARIO y T/C. Cambia
        # sólo: FACTURA = "N/A", CONCEPTO = "Propina", IMPORTE = monto de la
        # propina, IVA = 0, ISR = 0, ISH/IEPS = 0, RETENCIÓN = 0, TOTAL = la propina.
        # COBRADO/FACTURADO/DIFERENCIA quedan vacías (el renglón principal ya
        # contempla la propina en su fórmula de DIFERENCIA — no duplicamos).
        if propina_mxn > 0 or propina_ext > 0:
            ws.row_dimensions[row].height = 15
            propina_bg = 'F0FDF4'  # tinte verde muy sutil para agrupar con el padre

            # Pintar toda la banda primero (mismo fondo) para que las columnas
            # sin contenido (COBRADO/FACTURADO/DIFERENCIA) mantengan el tinte.
            for c in range(2, 24):
                pcell = ws.cell(row=row, column=c)
                pcell.fill = PatternFill('solid', start_color=propina_bg)
                pcell.border = Border(bottom=Side(style='hair', color=BORDER_LIGHT))

            # Valores de la propina, en paralelo al renglón principal (incluye
            # el caso USD: IMPORTE/T/C usan la celda editable de T/C).
            if is_usd_row and propina_ext > 0:
                p_importe_val = f'={propina_ext}*{tc_ref}'
                p_monto_usd   = propina_ext
                p_tc_val      = f'={tc_ref}'
            else:
                p_importe_val = propina_mxn
                p_monto_usd   = propina_ext if propina_ext > 0 else 0
                p_tc_val      = tipo_cambio

            # Concepto: para moneda extranjera incluye símbolo + monto nativo.
            if propina_ext > 0 and moneda_code != 'MXN':
                symbol = currency_symbol(moneda_code)
                concepto_prop = f'Propina {symbol}{propina_ext:,.2f} {moneda_code}'
            else:
                concepto_prop = 'Propina'

            pcells = [
                (2,  rfc_disp,               rfc_align, 'rfc'),
                (3,  str(g.get('proveedor', '') or '').upper(), 'left', 'normal_bold'),
                (4,  tipo,                   'center', 'badge_tipo'),
                (5,  poliza_row,             'center', 'normal'),
                (6,  'N/A',                  'center', 'normal'),
                (7,  fecha_fac,              'center', 'date'),
                (8,  fecha_cobro,            'center', 'date'),
                (9,  concepto_prop,          'left',   'normal'),
                (10, p_importe_val,          'center', 'currency'),
                (11, 0,                      'center', 'currency'),   # IVA
                (12, 0,                      'center', 'currency'),   # ISR
                (13, 0,                      'center', 'currency'),   # ISH/IEPS
                (14, 0,                      'center', 'currency'),   # RETENCIÓN
                (15, f'=J{row}+K{row}-L{row}+M{row}-N{row}', 'center', 'currency_bold'),  # TOTAL
                (16, forma,                  'center', 'badge_pago'),
                (17, banco,                  'center', 'normal'),
                (18, p_monto_usd,            'center', 'currency_ext'),
                (19, p_tc_val,               'center', 'tipocambio'),
                (20, usuario,                'center', 'normal'),
            ]
            for col, val, align, style_type in pcells:
                cell = ws.cell(row=row, column=col, value=val)
                cell.fill = PatternFill('solid', start_color=propina_bg)
                cell.border = Border(bottom=Side(style='hair', color=BORDER_LIGHT))
                cell.alignment = Alignment(
                    horizontal=align,
                    vertical='center',
                    indent=2 if align == 'left' else 0
                )
                style_data_cell(cell, style_type, tipo=tipo, diff_num=None,
                                moneda=moneda_code)

            # Outer spacers stay on page bg so the propina band fits inside
            # the table boundary like every other data row.
            ws.cell(row=row, column=1).fill = PatternFill('solid', start_color=BG_PAGE)
            ws.cell(row=row, column=24).fill = PatternFill('solid', start_color=BG_PAGE)

            row += 1

    # ═══ ADAPTIVE TOTALS — placed right after the last data row ═══
    ws.row_dimensions[row].height = 14  # spacer
    row += 1

    ws.row_dimensions[row].height = 32
    for c in range(2, 24):
        cell = ws.cell(row=row, column=c)
        cell.fill = PatternFill('solid', start_color=SMTO_BLACK)
        cell.border = Border()

    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
    lbl = ws.cell(row=row, column=2, value='TOTAL CUENTA')
    lbl.font = Font(name='Aptos', size=12, bold=True, color=WHITE)
    lbl.alignment = Alignment(horizontal='right', vertical='center', indent=2)
    lbl.fill = PatternFill('solid', start_color=SMTO_BLACK)

    # Same SUM formulas as the KPI cards above so the two views always agree.
    # `data_first` / `data_last` cubren filas principales + sub-filas de propina.
    # Tras agregar ISR (L) e ISH/IEPS (M): IMPORTE=J, IVA=K, ISR=L, ISH/IEPS=M,
    # RETENCIÓN=N, TOTAL=O, MONTO USD=R. T/C (S), BANCO (Q), USUARIO (T) no se
    # totalizan.
    totals = [
        (10, f'=SUM(J{data_first}:J{data_last})', False),  # IMPORTE
        (11, f'=SUM(K{data_first}:K{data_last})', False),  # IVA
        (12, f'=SUM(L{data_first}:L{data_last})', False),  # ISR (nueva)
        (13, f'=SUM(M{data_first}:M{data_last})', False),  # ISH/IEPS
        (14, f'=SUM(N{data_first}:N{data_last})', False),  # RETENCIÓN
        (15, f'=SUM(O{data_first}:O{data_last})', True),   # TOTAL
        (18, f'=SUM(R{data_first}:R{data_last})', False),  # MONTO M.E.
    ]
    for col, val, is_main in totals:
        cell = ws.cell(row=row, column=col, value=val)
        cell.number_format = '"$"#,##0.00'
        cell.font = Font(
            name='Aptos',
            size=16 if is_main else 11,
            bold=is_main,
            color=SMTO_GREEN if is_main else WHITE,
        )
        cell.alignment = Alignment(horizontal='right', vertical='center', indent=2)
        cell.fill = PatternFill('solid', start_color=SMTO_BLACK)

    # El total de MONTO M.E. sigue siendo la misma SUM(R) (numérica, para no
    # romper plantillas que leen la banda), pero el formato dice la verdad:
    # con UNA divisa lleva su símbolo (RM 480.00); con VARIAS va sin símbolo —
    # esa suma mezcla monedas y el equivalente real en MXN vive en la tarjeta
    # M. EXTRANJERA (MXN) de arriba.
    if len(monedas_ext) == 1:
        ws.cell(row=row, column=18).number_format = currency_format(monedas_ext[0])
    elif len(monedas_ext) > 1:
        ws.cell(row=row, column=18).number_format = '#,##0.00'

    # FORMA PAGO (P), BANCO (Q), T/C (S) y USUARIO (T) en la banda de totales
    # sólo llevan el fill negro (no son agregables).
    for c in (16, 17, 19, 20):
        ws.cell(row=row, column=c).fill = PatternFill('solid', start_color=SMTO_BLACK)

    # ── Totales de conciliación: COBRADO (U), FACTURADO incl. propinas (V),
    # DIFERENCIA (W) ── Se escriben como VALOR (no fórmula) por la misma razón
    # que las celdas por-renglón: renderizan en cualquier visor sin recálculo.
    total_diff_band = round(total_cobrado_band - total_facturado_band, 2)
    band_totals = [
        (21, round(total_cobrado_band, 2)),   # COBRADO
        (22, round(total_facturado_band, 2)),  # FACTURADO (incluye propinas)
    ]
    for col, val in band_totals:
        cell = ws.cell(row=row, column=col, value=val)
        cell.number_format = '"$"#,##0.00'
        cell.font = Font(name='Aptos', size=11, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill('solid', start_color=SMTO_BLACK)

    # DIFERENCIA total: color por signo (misma convención que la celda por
    # renglón — negativo rojo, positivo verde). ±5¢ se considera cuadrado y va
    # en blanco para no teñir de rojo un simple redondeo. Rojo claro para que
    # sea legible sobre el fondo negro de la banda.
    if total_diff_band < -0.05:
        diff_color = 'F87171'
    elif total_diff_band > 0.05:
        diff_color = SMTO_GREEN
    else:
        diff_color = WHITE
    dcell = ws.cell(row=row, column=23, value=total_diff_band)
    dcell.number_format = '"$"#,##0.00'
    dcell.font = Font(name='Aptos', size=11, bold=True, color=diff_color)
    dcell.alignment = Alignment(horizontal='center', vertical='center')
    dcell.fill = PatternFill('solid', start_color=SMTO_BLACK)

    # ═══ FOOTER — one spacer row + a right-aligned version line ═══
    row += 2  # blank spacer + footer row
    ws.row_dimensions[row].height = 18
    ws.merge_cells(start_row=row, start_column=11, end_row=row, end_column=23)
    ft = ws.cell(row=row, column=11)
    ft.value = 'SMTO Engineering · v8.96'
    ft.font = Font(name='Aptos', size=8, italic=True, color=TEXT_MUTED)
    ft.alignment = Alignment(horizontal='right', vertical='center')
    ft.fill = PatternFill('solid', start_color=BG_PAGE)

    # ═══ LOGO — cropped of padding, anchored at B1 next to the title ═══
    logo_path = find_logo()
    if logo_path:
        try:
            pil = PILImage.open(logo_path)
            pil = crop_logo(pil)
            target_h = 64
            ratio = target_h / pil.height
            target_w = int(pil.width * ratio)
            pil = pil.resize((target_w, target_h), PILImage.LANCZOS)
            with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp:
                logo_tmp = tmp.name
            pil.save(logo_tmp, 'PNG')
            img = XLImage(logo_tmp)
            img.width = target_w
            img.height = target_h
            img.anchor = 'B1'
            ws.add_image(img)
        except Exception as e:
            print(f'Logo: {e}')

    # Unlock cells so they're editable, PERO dejamos las celdas con fórmula
    # en locked=True (el default). Razón: Excel marca con triángulo verde
    # ("fórmula desprotegida") cualquier celda que tenga fórmula y esté
    # desbloqueada. Como la hoja NO está protegida, el estado locked no afecta
    # la edición (todo es editable igual), así que dejar las fórmulas locked
    # solo sirve para quitar esa advertencia y dejar el archivo más limpio.
    _unlocked = Protection(locked=False)
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith('='):
                continue  # fórmula → se queda locked=True, sin triángulo
            cell.protection = _unlocked

    return wb

class handler(BaseHTTPRequestHandler):
    def do_GET(self):
        diag = {'ok': IMPORT_ERROR is None, 'python_version': sys.version, 'logo_path': find_logo(), 'import_error': IMPORT_ERROR}
        self.send_response(200)
        self.send_header('Content-Type', 'application/json')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.end_headers()
        self.wfile.write(json.dumps(diag, indent=2).encode())

    def do_POST(self):
        try:
            if IMPORT_ERROR:
                raise RuntimeError(f'Import failed:\n{IMPORT_ERROR}')
            content_length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(content_length)
            data = json.loads(body)
            # Backward compatible: old clients send a bare array, new clients
            # send { gastos: [...], colaborador: '...' }.
            if isinstance(data, dict):
                gastos = data.get('gastos', [])
                colaborador = data.get('colaborador', '')
                poliza_numero = data.get('polizaNumero', 'N/A')
                polizas_map = data.get('polizasMap', {}) or {}
            else:
                gastos = data
                colaborador = ''
                poliza_numero = 'N/A'
                polizas_map = {}
            gastos = sanitize(gastos)
            colaborador = sanitize(colaborador)
            poliza_numero = sanitize(poliza_numero)
            polizas_map = sanitize(polizas_map) if isinstance(polizas_map, dict) else {}
            wb = build_workbook(gastos, colaborador, poliza_numero, polizas_map)
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                tmp_path = tmp.name
            wb.save(tmp_path)
            with open(tmp_path, 'rb') as f:
                file_data = f.read()
            os.unlink(tmp_path)
            self.send_response(200)
            self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition', 'attachment; filename="Reporte_Gastos_SMTO.xlsx"')
            self.send_header('Content-Length', str(len(file_data)))
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(file_data)
        except Exception as e:
            err = traceback.format_exc()
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(json.dumps({'error': str(e), 'trace': err}).encode())

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()
